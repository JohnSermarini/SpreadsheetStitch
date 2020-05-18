
#################################
#      Spreadsheet Stitch       
#        John Sermarini         
#            v0.9               
#################################

########## To Do List ###########
#
# - Create Readme
# - Clean imports to shrink .exe file size
# - Mac build set up
# - Add image dimensions to GUI
# - Improve color adjustment algorithm for larger images
# - Fix thread closing issue
# 
#################################

########## Wish List ############
#
# - Choose output folder/file name
# - Reimplement DMC Colors
# - Mac port
# - Custom symbols
# - Trim white space option
#
#################################


#import sys
from sys import argv as argv
from os import path, mkdir
from PIL import Image
from openpyxl import styles
from openpyxl import Workbook
from openpyxl import load_workbook
from string import ascii_uppercase
from numpy import rot90, fliplr
from numpy.linalg import norm
#import numpy as np
import tkinter as tk
import tkinter.font as tkFont
from tkinter import filedialog
from tkinter import messagebox
from tkinter.ttk import Progressbar
#import matplotlib as mpl
from matplotlib import pyplot as plt
from matplotlib import rcParams
from matplotlib.widgets import Slider as pltSlider
from matplotlib.widgets import Button as pltButton
from copy import deepcopy
from colorsys import rgb_to_hsv, hsv_to_rgb
from threading import Thread


## CSV formatting values
column_size = 2.8 # This number is about 20 pixels, same as the default height
cell_fill_type = 'solid'
legend_buffer = 1

## Program aesthetics values
error_box_header = "Error"
window_width = 30
window_title = "Spreadsheet Stitch"
color_base = "#217346"

## Program functionality values
file_path = ""
csv_output_directory = "Patterns"
max_color_input = 32
min_color_input = 1
max_dimension_input = 500
min_dimension_input = 1

## GUI values
label_file_selected = None
progress_bar = None
button_select_file = None
button_preview = None


def main(argv):
	## Init
	window = tk.Tk()
	guivar_checkbox_use_dmc = tk.IntVar()
	global label_file_selected
	global progress_bar
	global button_select_file
	global button_preview

	## Configure GUI
	window.title(window_title)
	set_window_icon(window)
	#window.geometry("400x300")
	window.resizable(width=False, height=False)
	window.configure(background="white")
	font = tkFont.Font(family="Times", size=13)
	## File select
	label_file_selected = tk.Label(window, font=font, width=window_width, text="No File Selected", fg="red")
	label_file_selected.pack(fill="both", expand=True)
	button_select_file = tk.Button(window, font=font, text="Select File", width=window_width, command=lambda : user_select_file())
	button_select_file.pack(fill="both", expand=True)
	## Width
	tk.Label(window, font=font, width=window_width, text="Width [" + str(min_dimension_input) + " - " + str(max_dimension_input) + "]").pack(fill="both", expand=True)
	entry_width = tk.Entry(window, font=font, width=window_width)
	entry_width.pack(fill="both", expand=True)
	## Height
	tk.Label(window, font=font, width=window_width, text="Height [" + str(min_dimension_input) + " - " + str(max_dimension_input) + "]").pack(fill="both", expand=True)
	entry_height = tk.Entry(window, font=font, width=window_width)
	entry_height.pack(fill="both", expand=True)
	## Colors
	tk.Label(window, font=font, width=window_width, text="Number of Colors [" + str(min_color_input) + " - " + str(max_color_input) + "]").pack(fill="both", expand=True)
	entry_num_colors = tk.Entry(window, font=font, width=window_width)
	entry_num_colors.pack(fill="both", expand=True)
	#checkbox_use_dmc_colors = tk.Checkbutton(text="Use DMC color palette", variable=guivar_checkbox_use_dmc, onvalue=True, offvalue=False)
	#checkbox_use_dmc_colors.pack()
	## Next
	button_preview = tk.Button(window, font=font, text="Next", width=window_width, command=lambda : show_preview(guivar_checkbox_use_dmc.get(), entry_width.get(), entry_height.get(), entry_num_colors.get()))
	button_preview.pack(fill="both", expand=True)
	## Progress bar
	progress_bar = Progressbar(window, orient=tk.HORIZONTAL, length=100, value=0, mode='determinate')
	progress_bar.pack(fill="both", expand=True)
	set_progress(0, 1)
	window.mainloop()


#Create output directory if it does not exist
def check_output_directory():
	if not path.isdir(csv_output_directory):
		mkdir(csv_output_directory)


def get_output_file_name(file_name):
	output_name = ""	
	file_split = file_name.split(".")
	for i in range(0, len(file_split) - 1):
		output_name = output_name + file_split[i]
	return output_name + ".xlsx"


def read_image(file_name):
	try:
		image = Image.open(file_name)
		return image
	except Exception as e:
		raise e
		#TODO image not found handling
		return None


def set_window_icon(window):
	try:
		#window.iconbitmap("graphics\\window_icon.ico")
		window.iconbitmap("graphics\\logo.ico")
	except Exception as e:
		print("Error: 'graphics\\logo.ico' not found")


#############################################
# IN: PIL image
# OUT: 2D array with each value containing a rgb tuple
# OUT: 2D array with each value containing an int representing the color map value
#############################################
def get_colors(image):
	used_colors = []
	colors = []
	color_map = []

	for x in range(0, image.size[0]): # Left column to right column
		column_colors = []
		column_map = []
		for y in range(0, image.size[1]): # Top row to bottom row
			pixel_color = image.getpixel((x,y))
			if(pixel_color not in used_colors):
				used_colors.append(pixel_color)
			pixel_map = used_colors.index(pixel_color)

			column_colors.append(pixel_color)
			column_map.append(pixel_map)
		colors.append(column_colors)
		color_map.append(column_map)

	return colors, color_map


def rgb_to_hex(color):
	# Note: Have color[3] for alpha for future expansion.
	return '%02x%02x%02x' % (color[0], color[1], color[2])


def get_cell_name(x, y):
	col = get_column(x + 1)
	row = get_row(y)

	return col + row


def get_column(num):
	# Lifted from : https://stackoverflow.com/questions/48983939/convert-a-number-to-excel-s-base-26
	# All credit to user 'poke'

	def divmod_excel(n):
	    a, b = divmod(n, 26)
	    if b == 0:
	        return a - 1, b + 26
	    return a, b

	chars = []
	while num > 0:
		num, d = divmod_excel(num)
		chars.append(ascii_uppercase[d - 1])
		#chars.append(string.ascii_uppercase[d - 1])
	return ''.join(reversed(chars)).upper()


def get_row(y):
	return str(y + 1)


def get_font_color(cell_color):
	r = cell_color[0]
	g = cell_color[1]
	b = cell_color[2]
	luma = 0.299*r + 0.587*g + 0.114*b
	luma = luma / 255.0 # Account for rgb scale being 0-255 instead of 0-1.0
	if luma > 0.7: # Cell is very bright
		return "00000000" # Black
	else: # Cell is very dark
		return "FFFFFFFF" # White



#############################################
# IN: pil image, int of amount of colors
# OUT: 2D array with each value containing a rgb tuple
#############################################
def reduce_color_palette(image, num_colors):
	#TODO use machine learning to do this instead? Current version kind of jank...

	pixel_image = image.convert("P", palette=Image.ADAPTIVE, colors=num_colors, dither=0)

	return pixel_image.convert("RGB") # convert back to RGB mode


#############################################
# DESC: Convert RGB colors to closest DMC color
# IN: 2D color array, 2D color mapping array, int of amount of colors
# OUT: 2D array with each value containing a rgb tuple
#############################################
def convert_colors_to_dmc(colors, color_map, num_colors):
	## Init color replacement array
	converted_colors = []
	for i in range(0, num_colors):
		converted_colors.append((-1, -1, -1))
	## Replace every color with the converted color
	for x in range(0, len(colors)):
		for y in range(0, len(colors[x])):
			map_value = color_map[x][y]
			## Converted color not set
			if(converted_colors[map_value] == (-1, -1, -1)):
				converted_colors[map_value] = find_closest_dmc_color(colors[x][y])
			## Replace color
			colors[x][y] = converted_colors[map_value]
	return colors


def find_closest_dmc_color(color):
	closest_distance = 1000000.0
	closest_index = -1

	dmc_colors = get_dmc_colors()

	for d in range(0, len(dmc_colors)):
		r = color[0] - dmc_colors[d][0]
		g = color[1] - dmc_colors[d][1]
		b = color[2] - dmc_colors[d][2]
		#euclidean_distance = np.linalg.norm([r, g, b])
		euclidean_distance = norm([r, g, b])
		if(euclidean_distance < closest_distance):
			closest_distance = euclidean_distance
			closest_index = d

	return dmc_colors[closest_index]


def get_dmc_colors():
	try:
		ws = load_workbook('color_chart.xlsx').worksheets[0]
		dmc_colors = []
		for row in ws.rows:
			r = row[2].value
			g = row[3].value
			b = row[4].value
			dmc_colors.append((r, g, b))
		return dmc_colors
	except Exception as e:
		tk.messagebox.showinfo(error_box_header, "Error: DMC color chart loading failed. Make sure 'color_chart.xlsx' is present and not open.")
		return None



def adjust_image_size(image, width, height):
	return image.resize((width, height))


def get_used_color_palette(colors, color_map):
	used_colors = []
	used_map = []

	## Get list of used colors
	for x in range(0, len(colors)):
		for y in range(0, len(colors[x])):
			color = colors[x][y]
			color_map_value = color_map[x][y]
			if color not in used_colors:
				used_colors.append(color)
				used_map.append(color_map_value)

	return used_colors, used_map


def get_dmc_name(use_dmc, color_rgb):
	if use_dmc:
		#TODO
		return "TODO"
	else:
		return "N/A"


def trim_image(image):
	# TODO
	# if row or col are entirely transparent... delete row/col
	return image


def get_worksheet_name():
	# TODO
	return "TODO"


def user_select_file():
	global file_path
	global label_file_selected
	## Get path
	file_path = filedialog.askopenfilename() # Returns string
	## Update label and console
	label_file_selected["text"] = get_file_name_from_path(file_path)
	label_file_selected["fg"] = "green"
	print("File:", file_path)


def create_color_grid(use_dmc, width, height, num_colors):
	global file_path

	# Check inputs for errors
	if not file_path_valid(file_path):
		return None, None
	if not dimensions_valid(width, height):
		return None, None
	if not num_colors_valid(num_colors):
		return None, None

	## Init
	file_name = get_file_name_from_path(file_path)
	width = int(width)
	height = int(height)
	num_colors = int(num_colors)

	## Get image from file
	image = read_image(file_path)
	image = adjust_image_size(image, width, height)
	image = trim_image(image)

	## Get colors from image
	image = reduce_color_palette(image, num_colors)
	colors, color_map = get_colors(image)
	if use_dmc:
		colors = convert_colors_to_dmc(colors, color_map, num_colors)

	## Close
	image.close()	

	return colors, color_map


'''#########################
		'Preview' GUI
 ---------------------------
|							|
|							|
|		Display Section		|
|			0.66%			|
|							|
|							|
|---------------------------|
|		Slider Section		|
|			0.33%			|
 ---------------------------
'''#########################


# Slider tutorial https://riptutorial.com/matplotlib/example/23577/interactive-controls-with-matplotlib-widgets
def show_preview(use_dmc, width, height, num_colors):
	## Init
	colors, color_map = create_color_grid(use_dmc, width, height, num_colors)
	if colors is None:
		return
	## Set preview window details
	rcParams['toolbar'] = "None"
	fig = plt.figure(num=window_title)
	## Set main GUI button status
	def handle_close(evt):
		enable_gui_buttons()
	disable_gui_buttons()
	fig.canvas.mpl_connect('close_event', handle_close)
	## Image
	ax_image = fig.add_axes([0, 0.33, 1.0, 0.66]) # add_axes([left, bottom, width, height])
	ax_image.imshow(get_preview_image_from_colors(colors), interpolation="none")
	ax_image.set_axis_off()
	## Add sliders
	sliders = []
	slider_width = 0.5
	slider_height = 0.05
	slider_left_offset = (1.0 - slider_width) / 2.0
	slider_vertical_buffer = 0.025
	## Brightness
	slider_vertical_offset = slider_vertical_buffer + slider_height + slider_vertical_buffer
	ax_slider = fig.add_axes([slider_left_offset, slider_vertical_offset, slider_width, slider_height]) # add_axes([left, bottom, width, height])
	slider_brightness = pltSlider(ax_slider, 'Brightness', 0, 2.0, valinit=1.0, valstep=0.01, color=color_base)
	sliders.append(slider_brightness)
	## Contrast
	slider_vertical_offset = slider_vertical_offset + slider_height + slider_vertical_buffer
	ax_contrast = fig.add_axes([slider_left_offset, slider_vertical_offset, slider_width, slider_height]) # add_axes([left, bottom, width, height])
	slider_contrast = pltSlider(ax_contrast, 'Contrast', 0, 1.0, valinit=0.5, valstep=0.01, color=color_base)
	sliders.append(slider_contrast)
	## Saturation
	slider_vertical_offset = slider_vertical_offset + slider_height + slider_vertical_buffer
	ax_saturation = fig.add_axes([slider_left_offset, slider_vertical_offset, slider_width, slider_height]) # add_axes([left, bottom, width, height])
	slider_saturation = pltSlider(ax_saturation, 'Saturation', 0, 1.0, valinit=1.0, valstep=0.01, color=color_base)
	sliders.append(slider_saturation)
	## Assign on changed update to sliders
	def adjust_colors_using_slider_vals(colors, brightness, contrast, saturation):
		new_colors = deepcopy(colors)
		new_colors = adjust_brightness(new_colors, brightness)
		new_colors = adjust_contrast(new_colors, contrast)
		new_colors = adjust_saturation(new_colors, saturation)
		return new_colors
	def update(val):
		# Reset image
		ax_image.clear()
		ax_image.set_axis_off()
		# Adjust colors
		new_colors = adjust_colors_using_slider_vals(colors, slider_brightness.val, slider_contrast.val, slider_saturation.val)
		# Update image using updated colors
		ax_image.imshow(get_preview_image_from_colors(new_colors), interpolation='none')
	for slider in sliders:
		slider.on_changed(update)
	## Reset Button
	def reset_sliders(event):
		for slider in sliders:
			slider.reset()
	ax_reset = fig.add_axes([0.25, slider_vertical_buffer, 0.25, slider_height])
	reset_button = pltButton(ax_reset, "Reset", hovercolor="0.75")
	reset_button.on_clicked(reset_sliders)
	## Commit Button
	def create(event):
		## Close the window
		plt.close()
		disable_gui_buttons() # Make sure buttons are disabled after they are re-enabled by the matplotlib window closing event
		## Create the workbook file
		new_colors = adjust_colors_using_slider_vals(colors, slider_brightness.val, slider_contrast.val, slider_saturation.val)
		#create_workbook(new_colors, color_map)
		workbook_thread = Thread(target=create_workbook, args=[new_colors, color_map])
		workbook_thread.daemon = True
		workbook_thread.start()
	ax_create = fig.add_axes([0.5, slider_vertical_buffer, 0.25, slider_height])
	create_button = pltButton(ax_create, "Create", hovercolor="0.75")
	create_button.on_clicked(create)
	## Show
	plt.show()


def get_preview_image_from_colors(colors):
	## Adjust orientation
	#rotated_colors = np.rot90(colors, k=3, axes=(0,1))
	rotated_colors = rot90(colors, k=3, axes=(0,1))
	#rotated_colors = np.fliplr(rotated_colors)
	rotated_colors = fliplr(rotated_colors)
	return rotated_colors


def adjust_brightness(colors, brightness):
	new_colors = []
	for x in range(0, len(colors)):
		new_colors.append(deepcopy(colors[x]))
		for y in range(0, len(colors[x])):
			r = min(255, int(colors[x][y][0] * brightness))
			g = min(255, int(colors[x][y][1] * brightness))
			b = min(255, int(colors[x][y][2] * brightness))
			new_colors[x][y] = (r, g, b)
	return new_colors


# Tutorial: https://www.dfstudios.co.uk/articles/programming/image-programming-algorithms/image-processing-algorithms-part-5-contrast-adjustment/
def adjust_contrast(colors, contrast):
	new_colors = []
	contrast = int(-128.0 + (256.0 * contrast)) # 0 = -128, 0.5 = 0, 1.0 = +128
	f = (259 * (contrast + 255)) / (255 * (259 - contrast))
	for x in range(0, len(colors)):
		new_colors.append(deepcopy(colors[x]))
		for y in range(0, len(colors[x])):
			r = max(0, min(255, f * (int(colors[x][y][0]) - 128) + 128))
			g = max(0, min(255, f * (int(colors[x][y][1]) - 128) + 128))
			b = max(0, min(255, f * (int(colors[x][y][2]) - 128) + 128))
			new_colors[x][y] = (r, g, b)
	return new_colors


def adjust_saturation(colors, saturation):
	## Convert to HSV, adjust S, convert back
	new_colors = []
	for x in range(0, len(colors)):
		new_colors.append(deepcopy(colors[x]))
		for y in range(0, len(colors[x])):
			## Convert RGB to HSV
			hsv = list(rgb_to_hsv(colors[x][y][0], colors[x][y][1], colors[x][y][2]))
			## Adjust saturation in HSV
			hsv[1] = hsv[1] * saturation
			## Convert adjusted HSV back to RGB
			rgb = hsv_to_rgb(hsv[0], hsv[1], hsv[2])
			new_colors[x][y] = (int(rgb[0]), int(rgb[1]), int(rgb[2]))
	return new_colors


def create_workbook(colors, color_map):
	global file_path
	## Init
	file_name = get_file_name_from_path(file_path)
	#colors, color_map = create_color_grid(use_dmc, width, height, num_colors)
	#if colors is None:
		#return
	## Create worksheet
	wb = Workbook()
	ws = wb.create_sheet(file_name, index=0)
	## Fill worksheet
	#fill_type = 'solid'
	for x in range(0, len(colors)):
		print("Converting - " +  str(x) + "/" + str(len(colors)) + " to Excel")
		set_progress(x + 1, len(colors))
		for y in range(0, len(colors[x])):
			cell_color = rgb_to_hex(colors[x][y])
			font_color = get_font_color(colors[x][y])
			#font_color = "FFFFFFFF" # White
			cell_symbol = color_map[x][y]
			cell_alignment = styles.Alignment(horizontal='center')
			cell_fill = styles.PatternFill(fill_type=cell_fill_type, start_color=cell_color, end_color=cell_color)
			cell_border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))
			cell_font = styles.Font(name='Calibri', bold=False, italic=False, color=font_color)
			cell_name = get_cell_name(x, y)
			ws[cell_name].alignment  = cell_alignment
			ws[cell_name].value = cell_symbol
			ws[cell_name].fill = cell_fill
			ws[cell_name].border = cell_border
			ws[cell_name].font = cell_font
		ws.column_dimensions[get_column(x + 1)].width = column_size # Set column size
	print("Conversion complete")
	## Add legend
	used_colors, used_map = get_used_color_palette(colors, color_map)
	#width = len(colors[0])
	width = len(colors)
	for c in range(-1, len(used_colors)):
		if(c == -1):
			ws[get_cell_name(width + legend_buffer, 0)].value = "Color"
			#ws[get_cell_name(width + legend_buffer + 1, 0)].value = "DMC Name"			
			#ws[get_cell_name(width + legend_buffer + 2, 0)].value = "HEX"
			#ws[get_cell_name(width + legend_buffer + 3, 0)].value = "Red Value"
			#ws[get_cell_name(width + legend_buffer + 4, 0)].value = "Green Value"
			#ws[get_cell_name(width + legend_buffer + 5, 0)].value = "Blue Value"
			ws[get_cell_name(width + legend_buffer + 1, 0)].value = "HEX"
			ws[get_cell_name(width + legend_buffer + 2, 0)].value = "Red Value"
			ws[get_cell_name(width + legend_buffer + 3, 0)].value = "Green Value"
			ws[get_cell_name(width + legend_buffer + 4, 0)].value = "Blue Value"
			continue		
		color_rgb = used_colors[c]
		color_symbol = used_map[c]
		color_hex = rgb_to_hex(color_rgb)
		font_color = get_font_color(color_rgb)
		cell_font = styles.Font(color=font_color)
		ws[get_cell_name(width + legend_buffer, c + 1)].fill = styles.PatternFill(fill_type=cell_fill_type, start_color=color_hex, end_color=color_hex)
		ws[get_cell_name(width + legend_buffer, c + 1)].value = str(color_symbol)
		ws[get_cell_name(width + legend_buffer, c + 1)].font = cell_font
		#ws[get_cell_name(width + legend_buffer + 1, c + 1)].value = get_dmc_name(use_dmc, color_rgb)
		#ws[get_cell_name(width + legend_buffer + 2, c + 1)].value = str(color_hex)
		#ws[get_cell_name(width + legend_buffer + 3, c + 1)].value = str(color_rgb[0])
		#ws[get_cell_name(width + legend_buffer + 4, c + 1)].value = str(color_rgb[1])
		#ws[get_cell_name(width + legend_buffer + 5, c + 1)].value = str(color_rgb[2])
		ws[get_cell_name(width + legend_buffer + 1, c + 1)].value = str(color_hex)
		ws[get_cell_name(width + legend_buffer + 2, c + 1)].value = str(color_rgb[0])
		ws[get_cell_name(width + legend_buffer + 3, c + 1)].value = str(color_rgb[1])
		ws[get_cell_name(width + legend_buffer + 4, c + 1)].value = str(color_rgb[2])
	## Save the file
	check_output_directory()
	output_directory = csv_output_directory
	output_file_name = get_output_file_name(file_name)
	output_file_path = output_directory + "\\" + output_file_name
	save_success = save_wb(wb, output_file_path)
	if save_success:
		print(output_file_name + " created")
		tk.messagebox.showinfo("Success", output_file_name + " created in folder '" + output_directory + "'")
	else:
		print(output_file_name + " save failed")
		tk.messagebox.showinfo(error_box_header, "Error: Save failed. Make sure file '" + get_file_name_from_path(output_file_name) + "' is not already open on computer.")
	set_progress(0, 1)
	enable_gui_buttons()


def set_progress(current_val, max_val):
	global progress_bar
	progress = (float(current_val) / float(max_val)) * 100
	progress_bar["value"] = progress


def enable_gui_buttons():
	global button_select_file
	global button_preview
	button_select_file["state"] = "normal"
	button_preview["state"] = "normal"


def disable_gui_buttons():
	global button_select_file
	global button_preview	
	button_select_file["state"] = "disable"
	button_preview["state"] = "disable"


#############################################
# DESC: Save the workbook as an excel file
# IN: workbook, file path
# OUT: boolean indicating success
#############################################
def save_wb(wb, output_file_path):
	try:
		wb.save(output_file_path)
		return True
	except Exception as e:
		return False


def file_path_valid(file_path):
	## Check for empty path
	if file_path == "":
		print("Error: Path file path empty.")
		messagebox.showinfo(error_box_header, "Error: Path file path empty.")
		return False
	## Check file type
	file_extension = file_path[-5:].lower()
	if  ".jpg" not in file_extension.lower() and \
		".png" not in file_extension.lower() and \
		".jpeg" not in file_extension.lower() :
		print("Error: File must be type '.png' or '.jpg'")
		messagebox.showinfo(error_box_header, "Error: File must be type '.png' or '.jpg'")
		return False
	return True


def dimensions_valid(width, height):
	## Check if height and width are numbers
	if not width.isnumeric():
		print("Error: Width contains non-numeric characters.")
		messagebox.showinfo(error_box_header, "Error: Width contains non-numeric characters.")
		return False
	if not height.isnumeric():
		print("Error: Height contains non-numeric characters.")
		messagebox.showinfo(error_box_header, "Error: Height contains non-numeric characters.")
		return False
	## Check if height in width are within the desired range
	width = int(width)
	height = int(height)
	if width < min_dimension_input or width > max_dimension_input:
		print("Error: Width '" + str(width) + "' not valid. Must be between " + str(min_dimension_input) + " and " + str(max_dimension_input) + ".")
		messagebox.showinfo(error_box_header, "Error: Width '" + str(width) + "' not valid. Must be between " + str(min_dimension_input) + " and " + str(max_dimension_input) + ".")
		return False
	if height < min_dimension_input or height > max_dimension_input:
		print("Error: Height '" + str(height) + "' not valid. Must be between " + str(min_dimension_input) + " and " + str(max_dimension_input) + ".")
		messagebox.showinfo(error_box_header, "Error: Height '" + str(height) + "' not valid. Must be " + str(min_dimension_input) + " and " + str(max_dimension_input) + ".")
		return False
	return True


def num_colors_valid(num_colors):
	## Check if number is numeric
	if not num_colors.isnumeric():
		print("Error: Number of Colors contains non-numeric characters.")
		messagebox.showinfo(error_box_header, "Error: Number of Colors contains non-numeric characters.")
		return False
	## Check if in desired range
	num_colors = int(num_colors)
	if num_colors < min_color_input or num_colors > max_color_input:
		print("Error: Number of Colors '" + str(num_colors) + "' not valid. Must be between "  + str(min_color_input) + " and " + str(max_color_input))
		messagebox.showinfo(error_box_header, "Error: Number of Colors '" + str(num_colors) + "' not valid. Must be between "  + str(min_color_input) + " and " + str(max_color_input))
		return False
	return True


def get_file_name_from_path(file_path):
	return file_path.split("/")[-1]


if __name__ == "__main__":
	main(argv[1:])
