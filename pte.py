"""
Name: TBD Cross Stich application 
Desc: ~
Auth: John Sermarini
"""

import sys
from PIL import Image
import os
from openpyxl import styles
from openpyxl import Workbook
from openpyxl import load_workbook
import string
import numpy as np
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
import matplotlib as mpl
from matplotlib import pyplot as plt


################################################################################################
# TODO check out openpyxl.utils.cell.get_column_letter(idx)
# from openpyxl.utils import get_column_letter

# TODO settings to GUI
# PYQT5????

# TODO add pixelizing to decrease image complexity

# TODO get dmc name

# TODO trim white space in image

# TODO give option for outer buffer in image after trimming

# TODO saturation slider?

# TODO custom symbols/colors

"""
TODO error checks:
Is file open in excel during save.
is path less than 5 characters
Pic has too many colors?????
check Get_file_name_from_path is only a filename and not a path
jpeg having two .'s
add error check for having the dmc color_chart file open
"""
################################################################################################


# CSV formatting values
column_size = 2.8 # This number is about 20 pixels, same as the default height
cell_fill_type = 'solid'
legend_buffer = 1

# Program aesthetics values
error_box_header = "Error"

# Program functionality values
file_path = ""

# GUI values
label_file_selected = None


def main(argv):
	# Init
	window = tk.Tk()
	guivar_checkbox_use_dmc = tk.IntVar()
	global label_file_selected

	# Configure GUI
	window.title("CSX")
	window.geometry("300x300")
	window.configure(background="white")
	set_window_icon(window)
	tk.Label(window, text="Width [1 - 99]:").pack()
	entry_width = tk.Entry(window, width=300)
	entry_width.pack()
	tk.Label(window, text="Height [1 - 99]:").pack()
	entry_height = tk.Entry(window, width=300)
	entry_height.pack()
	tk.Label(window, text="Number of Colors [4 - 16]:").pack()
	entry_num_colors = tk.Entry(window, width=300)
	entry_num_colors.pack()
	checkbox_use_dmc_colors = tk.Checkbutton(text="Use DMC color palette", variable=guivar_checkbox_use_dmc, onvalue=True, offvalue=False)
	checkbox_use_dmc_colors.pack()
	button_select_file = tk.Button(window, text="Select File", width=300, command=lambda : user_select_file())
	button_select_file.pack()
	label_file_selected = tk.Label(window, width=300, text="No File Selected", fg="red")
	label_file_selected.pack()
	label_file_selected.pack()
	button_preview = tk.Button(window, text="Preview", width=300, command=lambda : show_preview(guivar_checkbox_use_dmc.get(), entry_width.get(), entry_height.get(), entry_num_colors.get()))
	button_preview.pack()
	button_create = tk.Button(window, text="Create", width=300, command=lambda : create_workbook(guivar_checkbox_use_dmc.get(), entry_width.get(), entry_height.get(), entry_num_colors.get()))
	button_create.pack()
	window.mainloop()


def get_output_directory():
	#TODO 
	#TODO make directory if it doesn't exist
	return "out"


def get_output_file_name(file_name):
	file_name = file_name[:-4]

	return file_name + ".xlsx"


def read_image(file_name):
	try:
		image = Image.open(file_name)

		return image
	except Exception as e:
		raise e
		#TODO image not found handling

		return None


def set_window_icon(window):
	try:
		window.iconbitmap("window_icon.ico")
	except Exception as e:
		print("Error: 'window_icon.ico' not found")


#############################################
# IN: PIL image
# OUT: 2D array with each value containing a rgb tuple
# OUT: 2D array with each value containing an int representing the color map value
#############################################
def get_colors(image):
	used_colors = []
	colors = []
	color_map = []

	for x in range(0, image.size[0]): # Left column to right column
		column_colors = []
		column_map = []
		for y in range(0, image.size[1]): # Top row to bottom row
			pixel_color = image.getpixel((x,y))
			if(pixel_color not in used_colors):
				used_colors.append(pixel_color)
			pixel_map = used_colors.index(pixel_color)

			column_colors.append(pixel_color)
			column_map.append(pixel_map)
		colors.append(column_colors)
		color_map.append(column_map)

	return colors, color_map


def rgb_to_hex(color):
	# Note: Have color[3] for alpha for future expansion.
	return '%02x%02x%02x' % (color[0], color[1], color[2])


def get_cell_name(x, y):
	col = get_column(x + 1)
	row = get_row(y)

	return col + row


def get_column(num):
	# Lifted from : https://stackoverflow.com/questions/48983939/convert-a-number-to-excel-s-base-26
	# All credit to user 'poke'

	def divmod_excel(n):
	    a, b = divmod(n, 26)
	    if b == 0:
	        return a - 1, b + 26
	    return a, b

	chars = []
	while num > 0:
		num, d = divmod_excel(num)
		chars.append(string.ascii_uppercase[d - 1])
	return ''.join(reversed(chars)).upper()


def get_row(y):
	return str(y + 1)


def get_font_color(cell_color):
	r = cell_color[0]
	g = cell_color[1]
	b = cell_color[2]
	luma = 0.299*r + 0.587*g + 0.114*b
	if luma < 0.1:
		return "FFFFFFFF" # White
	else:
		return "00000000" # Black



#############################################
# IN: pil image, int of amount of colors
# OUT: 2D array with each value containing a rgb tuple
#############################################
def reduce_color_palette(image, num_colors):
	#TODO use machine learning to do this instead? Current version kind of jank...

	pixel_image = image.convert("P", palette=Image.ADAPTIVE, colors=num_colors, dither=0)

	return pixel_image.convert("RGB") # convert back to RGB mode


#############################################
# DESC: Convert RGB colors to closest DMC color
# IN: 2D color array, 2D color mapping array, int of amount of colors
# OUT: 2D array with each value containing a rgb tuple
#############################################
def convert_colors_to_dmc(colors, color_map, num_colors):
	#converted_colors = np.full(len(colors), -1, dtype=object)
	converted_colors = []
	for i in range(0, num_colors):
		converted_colors.append((-1, -1, -1))

	for x in range(0, len(colors)):
		print("Converting - " +  str(x) + "/" + str(len(colors)) + " to DMC color palette")
		for y in range(0, len(colors[x])):
			map_value = color_map[x][y]
			if(converted_colors[map_value] == (-1, -1, -1)): # converted color not set
				converted_colors[map_value] = find_closest_dmc_color(colors[x][y])
				print("Converted " + str(colors[x][y]) + " to " + str(converted_colors[map_value]))
			colors[x][y] = converted_colors[map_value]

	return colors


def find_closest_dmc_color(color):
	closest_distance = 1000000.0
	closest_index = -1

	dmc_colors = get_dmc_colors()

	for d in range(0, len(dmc_colors)):
		r = color[0] - dmc_colors[d][0]
		g = color[1] - dmc_colors[d][1]
		b = color[2] - dmc_colors[d][2]
		euclidean_distance = np.linalg.norm([r, g, b])
		if(euclidean_distance < closest_distance):
			closest_distance = euclidean_distance
			closest_index = d

	return dmc_colors[closest_index]


def get_dmc_colors():
	#TODO error handling
	ws = load_workbook('color_chart.xlsx').worksheets[0]
	dmc_colors = []
	for row in ws.rows:
		r = row[2].value
		g = row[3].value
		b = row[4].value
		dmc_colors.append((r, g, b))

	return dmc_colors


def adjust_image_size(image, width, height):
	return image.resize((width, height))


def get_used_color_palette(colors, color_map):
	used_colors = []
	used_map = []

	# Get list of used colors
	for x in range(0, len(colors)):
		for y in range(0, len(colors[x])):
			color = colors[x][y]
			color_map_value = color_map[x][y]
			if color not in used_colors:
				used_colors.append(color)
				used_map.append(color_map_value)

	return used_colors, used_map


def get_dmc_name(use_dmc, color_rgb):
	if use_dmc:
		#TODO
		return "TODO"
	else:
		return "N/A"


def trim_image(image):
	# TODO
	# if row or col are entirely transparent... delete row/col
	return image


def get_worksheet_name():
	# TODO
	return "TODO"


def user_select_file():
	global file_path
	global label_file_selected
	# Get path
	file_path = filedialog.askopenfilename() # Returns string
	# Update label and console
	label_file_selected["text"] = get_file_name_from_path(file_path)
	label_file_selected["fg"] = "green"
	print("File:", file_path)


def create_color_grid(use_dmc, width, height, num_colors):
	global file_path

	# Check inputs for errors
	if not file_path_valid(file_path):
		return None, None
	if not dimensions_valid(width, height):
		return None, None
	if not num_colors_valid(num_colors):
		return None, None

	# Init
	file_name = get_file_name_from_path(file_path)
	width = int(width)
	height = int(height)
	num_colors = int(num_colors)

	# Get image from file
	image = read_image(file_path)
	image = adjust_image_size(image, width, height)
	image = trim_image(image)

	# Get colors from image
	image = reduce_color_palette(image, num_colors)
	colors, color_map = get_colors(image)
	if use_dmc:
		colors = convert_colors_to_dmc(colors, color_map, num_colors)

	# Close
	image.close()	

	return colors, color_map


def show_preview(use_dmc, width, height, num_colors):
	# Init
	colors, color_map = create_color_grid(use_dmc, width, height, num_colors)
	if colors is None:
		return
	# Set preview window details
	mpl.rcParams['toolbar'] = 'None'
	plt.figure(num='Preview')
	plt.axis('off')
	# Adjust orientation
	rotated_colors = np.rot90(colors, k=3, axes=(0,1))
	rotated_colors = np.fliplr(rotated_colors)
	# Display
	plt.imshow(rotated_colors, interpolation='none')
	plt.show()


def create_workbook(use_dmc, width, height, num_colors):
	global file_path
	# Init
	file_name = get_file_name_from_path(file_path)
	colors, color_map = create_color_grid(use_dmc, width, height, num_colors)
	if colors is None:
		return
	# Create worksheet
	wb = Workbook()
	ws = wb.create_sheet(file_name, index=0)
	# Fill worksheet
	#fill_type = 'solid'
	for x in range(0, len(colors)):
		print("Converting - " +  str(x) + "/" + str(len(colors)) + " to Excel")
		for y in range(0, len(colors[x])):
			cell_color = rgb_to_hex(colors[x][y])
			#font_color = get_font_color(colors[x][y])
			font_color = "FFFFFFFF" # White
			cell_symbol = color_map[x][y]
			cell_alignment = styles.Alignment(horizontal='center')
			cell_fill = styles.PatternFill(fill_type=cell_fill_type, start_color=cell_color, end_color=cell_color)
			cell_border = styles.Border(left=styles.Side(style='thin'), right=styles.Side(style='thin'), top=styles.Side(style='thin'), bottom=styles.Side(style='thin'))
			cell_font = styles.Font(name='Calibri', bold=False, italic=False, color=font_color)
			cell_name = get_cell_name(x, y)
			ws[cell_name].alignment  = cell_alignment
			ws[cell_name].value = cell_symbol
			ws[cell_name].fill = cell_fill
			ws[cell_name].border = cell_border
			ws[cell_name].font = cell_font
		ws.column_dimensions[get_column(x + 1)].width = column_size # Set column size
	print("Conversion complete")
	# Add legend
	used_colors, used_map = get_used_color_palette(colors, color_map)
	width = len(colors[0])
	for c in range(-1, len(used_colors)):
		if(c == -1):
			ws[get_cell_name(width + legend_buffer, 0)].value = "Color"
			ws[get_cell_name(width + legend_buffer + 1, 0)].value = "DMC Name"			
			ws[get_cell_name(width + legend_buffer + 2, 0)].value = "HEX"
			ws[get_cell_name(width + legend_buffer + 3, 0)].value = "Red Value"
			ws[get_cell_name(width + legend_buffer + 4, 0)].value = "Green Value"
			ws[get_cell_name(width + legend_buffer + 5, 0)].value = "Blue Value"
			continue		
		color_rgb = used_colors[c]
		color_symbol = used_map[c]
		color_hex = rgb_to_hex(color_rgb)
		ws[get_cell_name(width + legend_buffer, c + 1)].fill = styles.PatternFill(fill_type=cell_fill_type, start_color=color_hex, end_color=color_hex)
		ws[get_cell_name(width + legend_buffer, c + 1)].value = str(color_symbol)
		ws[get_cell_name(width + legend_buffer + 1, c + 1)].value = get_dmc_name(use_dmc, color_rgb)
		ws[get_cell_name(width + legend_buffer + 2, c + 1)].value = str(color_hex)
		ws[get_cell_name(width + legend_buffer + 3, c + 1)].value = str(color_rgb[0])
		ws[get_cell_name(width + legend_buffer + 4, c + 1)].value = str(color_rgb[1])
		ws[get_cell_name(width + legend_buffer + 5, c + 1)].value = str(color_rgb[2])
	# Save the file
	output_directory = get_output_directory()
	output_file_name = get_output_file_name(file_name)
	output_file_path = output_directory + "\\" + output_file_name
	save_success = save_wb(wb, output_file_path)
	if save_success:
		print(output_file_name + " created")
		messagebox.showinfo("Success", output_file_name + " created in folder '" + output_directory + "'")
	else:
		print(output_file_name + " save failed")
		messagebox.showinfo(error_box_header, "Error: Save failed. Make sure file '" + get_file_name_from_path(output_file_name) + "' is not already open on computer.")


#############################################
# DESC: Save the workbook as an excel file
# IN: workbook, file path
# OUT: boolean indicating success
#############################################
def save_wb(wb, output_file_path):
	try:
		wb.save(output_file_path)
		return True
	except Exception as e:
		return False


def file_path_valid(file_path):
	# Check for empty path
	if file_path == "":
		print("Error: Path file path empty.")
		messagebox.showinfo(error_box_header, "Error: Path file path empty.")
		return False
	# Check file type
	file_extension = file_path[-5:].lower()
	if  ".jpg" not in file_extension.lower() and \
		".png" not in file_extension.lower() and \
		".jpeg" not in file_extension.lower() :
		print("Error: File must be type '.png' or '.jpg'")
		messagebox.showinfo(error_box_header, "Error: File must be type '.png' or '.jpg'")
		return False
	return True


def dimensions_valid(width, height):
	# Check if eheight and width are numbers
	if not width.isnumeric():
		print("Error: Width contains non-numeric characters.")
		messagebox.showinfo(error_box_header, "Error: Width contains non-numeric characters.")
		return False
	if not height.isnumeric():
		print("Error: Height contains non-numeric characters.")
		messagebox.showinfo(error_box_header, "Error: Height contains non-numeric characters.")
		return False
	# Check if height in width are within the desired range
	width = int(width)
	height = int(height)
	if width < 1 or width > 99:
		print("Error: Width '" + str(width) + "' not valid. Must be between 1 and 99.")
		messagebox.showinfo(error_box_header, "Error: Width '" + str(width) + "' not valid. Must be between 1 and 99.")
		return False
	if height < 1 or height > 99:
		print("Error: Height '" + str(height) + "' not valid. Must be between 1 and 99.")
		messagebox.showinfo(error_box_header, "Error: Height '" + str(height) + "' not valid. Must be between 1 and 99.")
		return False
	return True


def num_colors_valid(num_colors):
	if not num_colors.isnumeric():
		print("Error: Number of Colors contains non-numeric characters.")
		messagebox.showinfo(error_box_header, "Error: Number of Colors contains non-numeric characters.")
		return False
	num_colors = int(num_colors)
	if num_colors < 4 or num_colors > 16:
		print("Error: Number of Colors '" + str(num_colors) + "' not valid. Must be between 4 and 16.")
		messagebox.showinfo(error_box_header, "Error: Number of Colors '" + str(num_colors) + "' not valid. Must be between 4 and 16.")
		return False
	return True


def get_file_name_from_path(file_path):
	return file_path.split("/")[-1]


if __name__ == "__main__":
	main(sys.argv[1:])